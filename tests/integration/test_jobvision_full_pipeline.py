"""
Full end-to-end integration test for the JobVision pipeline (Commit 5):

    JobVisionSpider (real) -> JobVisionParser (real) ->
    JobVisionNormalizer (real) -> JobValidator (real) ->
    JobIngestionPipeline (real) -> Repositories (real) ->
    PostgreSQL (real) -> JSONExporter (real) -> ExcelExporter (real)

The only thing faked is the `Downloader.fetch_html` transport itself — it
serves real, previously-captured JobVision HTML (the same fixtures
`tests/unit/test_jobvision.py` unit-tests in isolation) instead of making
an actual network/browser call, since this sandbox cannot reach
jobvision.ir. Every other component in this chain is the real, production
class.

This test intentionally drives two different real jobs through two
different code paths:
- Job 1 (id 1409285): its `raw_company_url` points at a company page we
  DO have a fixture for (`company_profile.html`, id 17161) — exercises
  the full company-page-crawl path in `Pipeline._resolve_company`.
- Job 2 (id 1434775): its `raw_company_url` points at a DIFFERENT company
  (id 572) we do NOT have a fixture for — the fake downloader raises
  `TransientCrawlError` for it, exercising `_resolve_company`'s graceful
  fallback to a name-only `Company` row built from `raw_company_name`.

It also exercises Phase 3.5's bonus-jobs-via-company-page path: fixture
company 17161's page (`company_profile.html`) embeds its own 10-job
`GetListOfCompanyJobPosts` list (job 1409285 plus 9 others, mostly
expired). None of those 9 extra job IDs have their own detail-page
fixture, so every one of them is expected to fall back to summary-only
data — this is the normal/expected path for expired postings on the real
site, not a failure.
"""

from __future__ import annotations

import uuid
from pathlib import Path

import openpyxl
import pytest

from app.core.exceptions import TransientCrawlError
from app.database.session import get_db_session
from app.dto.company_export_dto import CompanyExportDTO
from app.dto.job_export_dto import JobExportDTO
from app.exporters.excel_exporter import ExcelExporter
from app.exporters.json_exporter import JSONExporter
from app.models.enums import ScrapeStatus
from app.normalizers.jobvision_normalizer import JobVisionNormalizer
from app.parsers.jobvision_parser import JobVisionParser
from app.pipelines.job_pipeline import JobIngestionPipeline
from app.repositories import CompanyRepository, JobRepository, SourceRepository
from app.spiders.downloader import Downloader
from app.spiders.jobvision_spider import JobVisionSpider

pytestmark = pytest.mark.asyncio

FIXTURES_DIR = Path(__file__).parent.parent / "fixtures" / "jobvision"

_SYNTHETIC_LISTING_HTML = """
<html><body>
  <job-card><a class="jvt-relative desktop-job-card" href="/jobs/1409285/x"></a></job-card>
  <job-card><a class="jvt-relative desktop-job-card" href="/jobs/1434775/x"></a></job-card>
</body></html>
"""


class _FixtureDownloader(Downloader):
    """Serves real, captured JobVision HTML instead of a live browser."""

    def __init__(self, listing_url: str) -> None:
        self._retry_count = 0
        self._listing_url = listing_url
        self._job_html = {
            "/jobs/1409285/": (FIXTURES_DIR / "job_detail_1.html").read_text(encoding="utf-8"),
            "/jobs/1434775/": (FIXTURES_DIR / "job_detail_2.html").read_text(encoding="utf-8"),
        }
        self._company_html = {
            "/companies/17161/": (
                FIXTURES_DIR / "company_profile.html"
            ).read_text(encoding="utf-8"),
        }

    async def fetch_html(self, url: str) -> str:  # type: ignore[override]
        if url == self._listing_url:
            return _SYNTHETIC_LISTING_HTML
        for path_fragment, html in {**self._job_html, **self._company_html}.items():
            if path_fragment in url:
                return html
        # Company 572 (and anything else unmapped) lands here — simulates
        # a real crawl failure so `_resolve_company`'s graceful fallback
        # to a name-only company gets exercised for real.
        raise TransientCrawlError(f"No fixture available for {url}")


async def test_jobvision_pipeline_end_to_end_then_export():
    # Unique per invocation — see test_job_pipeline.py / test_repositories.py
    # for the same rationale: this runs against a real, persistent
    # PostgreSQL instance, so reusing the literal "jobvision" source code
    # across repeated test runs would find last run's jobs as "already
    # known" and turn every `jobs_created` assertion into `jobs_updated`.
    # `JobVisionSpider.SITE_CODE` is overridden per-instance (the parser's
    # `raw.source_code` field staying "jobvision" doesn't matter — it's
    # informational only, not part of any dedup key).
    site_code = f"jobvision_test_{uuid.uuid4().hex[:8]}"

    keyword = "برنامه نویس"
    spider_for_url = JobVisionSpider(downloader=None, max_pages=1, keyword=keyword)
    listing_url = spider_for_url.build_listing_url(1)
    downloader = _FixtureDownloader(listing_url)
    spider = JobVisionSpider(downloader, max_pages=1, keyword=keyword)
    spider.SITE_CODE = site_code

    async with get_db_session() as session:
        pipeline = JobIngestionPipeline(
            session,
            spider=spider,
            parser=JobVisionParser(),
            normalizer=JobVisionNormalizer(),
        )
        run = await pipeline.run()

        # 2 jobs from the direct keyword-search crawl + 9 bonus jobs
        # discovered via company 17161's own embedded job list (see module
        # docstring) = 11 total.
        assert run.jobs_found == 11
        assert run.jobs_created == 11
        assert run.status == ScrapeStatus.SUCCESS  # company fallback doesn't count as a job error

        source_repo = SourceRepository(session)
        source = await source_repo.get_by_code(site_code)
        assert source is not None

        job_repo = JobRepository(session)
        job1 = await job_repo.get_by_source_and_website_id(source.id, "1409285")
        job2 = await job_repo.get_by_source_and_website_id(source.id, "1434775")
        assert job1.title == "برنامه‌نویس - خانم"
        assert job2.title == "برنامه‌نویس فرانت‌اند (Front-end Developer)"

        # A bonus job discovered only through company 17161's own job list,
        # whose own detail page has no fixture — must have landed using
        # the graceful summary-only fallback (title present, but no
        # description since that field only exists on the real detail page).
        bonus_job = await job_repo.get_by_source_and_website_id(source.id, "1130943")
        assert bonus_job is not None
        assert bonus_job.title == "کارشناس استقرار و پشتیبانی نرم افزار (ERP)"
        assert bonus_job.description is None

        company_repo = CompanyRepository(session)
        company1 = await company_repo.get_by_source_and_website_id(source.id, "17161")
        assert company1 is not None
        assert company1.industry is not None  # full profile crawl succeeded -> has industry

        # Job 2's company: fixture-less company 572 -> graceful name-only
        # fallback, looked up by (source_id, name) since it has no
        # website_company_id in this path.
        company2 = await company_repo.get_by_source_and_name(source.id, "شرکت ارتباطات مبین نت")
        assert company2 is not None
        assert company2.industry is None  # name-only fallback, no profile data

        # --- Export layer: real JSONExporter/ExcelExporter, real files ---
        jobs_orm = await job_repo.list_by_source_with_relations(source.id)
        companies_orm = await company_repo.list_by_source_with_relations(source.id)
        job_dtos = [JobExportDTO.from_orm_job(j) for j in jobs_orm]
        company_dtos = [CompanyExportDTO.from_orm_company(c) for c in companies_orm]

        assert len(job_dtos) == 11
        # All 11 jobs belong to just the 2 companies from the direct-crawl
        # path — every bonus job is from company 17161, already resolved.
        assert len(company_dtos) == 2
        job1_export = next(d for d in job_dtos if d.website_job_id == "1409285")
        assert job1_export.company_name == "فرا سامانه//همکاران سیستم"

    tmp_json_dir = Path("/tmp/jvt_export_test/json")
    tmp_excel_dir = Path("/tmp/jvt_export_test/excel")

    json_path = JSONExporter(tmp_json_dir).export(job_dtos, company_dtos, source_code="jobvision")
    excel_path = ExcelExporter(tmp_excel_dir).export(
        job_dtos, company_dtos, source_code="jobvision"
    )

    assert json_path.exists()
    json_text = json_path.read_text(encoding="utf-8")
    assert "برنامه‌نویس - خانم" in json_text
    assert '"job_count": 11' in json_text

    assert excel_path.exists()
    workbook = openpyxl.load_workbook(excel_path)
    assert set(workbook.sheetnames) == {"Jobs", "Companies"}
    jobs_sheet = workbook["Jobs"]
    # header row + 11 data rows
    assert jobs_sheet.max_row == 12
    companies_sheet = workbook["Companies"]
    assert companies_sheet.max_row == 3
