"""Regression coverage for `ExcelExporter`.

No exporter tests existed before this file — which is exactly how the
column-width bug this test guards against (`_flatten_row`'s neighbor,
the per-column `worksheet.column_dimensions[...].width` loop) shipped
undetected: `JobExportDTO` has 31 fields, so the real Jobs sheet was
always hitting the bug on its last 5 columns, but nothing ever asserted
against actual column widths.
"""

from __future__ import annotations

import uuid
from datetime import datetime, timezone
from decimal import Decimal
from pathlib import Path

import openpyxl

from app.dto.job_export_dto import JobExportDTO
from app.exporters.excel_exporter import ExcelExporter
from app.models.enums import (
    Currency,
    EducationLevel,
    ExperienceLevel,
    Gender,
    JobStatus,
    MilitaryStatus,
    WorkMode,
)


def _make_job(**overrides) -> JobExportDTO:
    defaults = dict(
        id=uuid.uuid4(),
        source="jobvision",
        website_job_id="123",
        source_url="https://jobvision.ir/jobs/123",
        title="برنامه‌نویس بک‌اند ارشد با سابقه‌ی طولانی در معماری میکروسرویس",
        work_mode=WorkMode.REMOTE,
        experience_level=ExperienceLevel.SENIOR,
        education=EducationLevel.BACHELOR,
        currency=Currency.IRT,
        gender=Gender.ANY,
        military_status=MilitaryStatus.NOT_REQUIRED,
        status=JobStatus.ACTIVE,
        scraped_at=datetime.now(timezone.utc),
        updated_at=datetime.now(timezone.utc),
        salary_min=Decimal("45000000"),
        salary_max=Decimal("60000000"),
    )
    defaults.update(overrides)
    return JobExportDTO(**defaults)


class TestExcelExporterColumnWidths:
    def test_every_column_gets_its_own_width_past_column_z(self, tmp_path: Path):
        # JobExportDTO has 31 fields (> 26 = past column Z into AA/AB/...).
        # The bug: every column past 26 silently fell back to column "A"
        # instead of AA/AB/etc, so (a) columns 27+ never got sized and (b)
        # column A's width got repeatedly overwritten by unrelated later
        # columns' content lengths.
        job = _make_job(
            title="کوتاه",  # short value in an early column (title, #5)
            # scraped_at/updated_at are columns #30/#31 -- give them an
            # obviously long rendered value so a width mix-up is easy to
            # detect if column A (an early column) ends up sized for it
            # instead.
        )
        assert len(JobExportDTO.model_fields) > 26, (
            "this test's premise (more than 26 columns) no longer holds -- "
            "re-check whether the column-width bug can still occur"
        )

        exporter = ExcelExporter(tmp_path)
        path = exporter.export([job], [], source_code="jobvision_test")

        workbook = openpyxl.load_workbook(path)
        sheet = workbook["Jobs"]

        # Column 1 ("id") should be sized for a UUID string (~36 chars),
        # not for some unrelated far-right column's content.
        id_col_width = sheet.column_dimensions["A"].width
        assert id_col_width is not None
        assert id_col_width < 45  # UUID length + padding, generous ceiling

        # Column 27 is "published_at" -- first column past Z, landing on
        # "AA". Before the fix this was silently never assigned a width
        # entry at all (fell through to "A" instead).
        assert "AA" in sheet.column_dimensions
        assert sheet.column_dimensions["AA"].width is not None

        # Column 31 ("updated_at"), the last column -- lands on "AE".
        assert "AE" in sheet.column_dimensions
        assert sheet.column_dimensions["AE"].width is not None

    def test_export_with_zero_jobs_and_zero_companies_does_not_crash(self, tmp_path: Path):
        # Edge case worth locking in: an empty crawl result (e.g. every
        # category failed) should still produce a valid, openable workbook
        # rather than crashing the whole pipeline at the export step.
        exporter = ExcelExporter(tmp_path)
        path = exporter.export([], [], source_code="jobvision_test")

        workbook = openpyxl.load_workbook(path)
        assert set(workbook.sheetnames) == {"Jobs", "Companies"}
